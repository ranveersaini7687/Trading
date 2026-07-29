#!/usr/bin/env python3
"""
Paper Trader — Equity cash simulation on Long Buildup scanner signals.

Capital  : ₹10,00,000
Sizing   : ₹1,00,000 per trade  (max 10 positions)
Stop-loss: -1%  from entry
Target   : +2%  from entry
Exit only: SL hit or target hit

Run daily:
  python3 long_buildup_scanner.py   → refresh signals
  python3 paper_trader.py           → update P&L + Excel report
"""

import json
import os
import time
import pandas as pd
from datetime import datetime
from angel_api import fetch_ltps
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Live trading integration
try:
    from live_trader import get_live_trader
    LIVE_TRADER = get_live_trader()
except Exception as e:
    log_msg = f"[paper_trader] Warning: Live trading module unavailable: {e}"
    print(log_msg, flush=True)
    LIVE_TRADER = None

# ── Config ────────────────────────────────────────────────────────────────────
PORTFOLIO_FILE  = "paper_portfolio.json"
CONFIRM_PORTFOLIO_FILE = "paper_portfolio_confirm.json"
SCAN_FILE       = "scan_results.json"
EXCEL_FILE      = "trade_log.xlsx"
CONFIRM_LOG_FILE = "eod_confirm_log.json"

TOTAL_CAPITAL   = 1_000_000      # ₹10,00,000
MAX_POSITIONS   = 10              # up to 10 slots × ₹1L
ALLOC_PER_TRADE = 100_000         # ₹1,00,000 per trade
STOP_LOSS_PCT   = 1.0             # -1%
TARGET_PCT      = 2.0             # +2%
MAX_PER_SECTOR  = 1               # max 1 stock per sector group

# Broad sector groupings for concentration check
SECTOR_MAP = {
    "AUBANK":"BANK", "KOTAKBANK":"BANK", "HDFCBANK":"BANK", "ICICIBANK":"BANK",
    "SBIN":"BANK", "BANKBARODA":"BANK", "BANKINDIA":"BANK", "PNB":"BANK",
    "CANBK":"BANK", "FEDERALBNK":"BANK", "BANDHANBNK":"BANK", "INDUSINDBK":"BANK",
    "YESBANK":"BANK", "RBLBANK":"BANK", "IDFCFIRSTB":"BANK",
    "ABCAPITAL":"NBFC", "SHRIRAMFIN":"NBFC", "CHOLAFIN":"NBFC", "BAJFINANCE":"NBFC",
    "BAJAJFINSV":"NBFC", "MANAPPURAM":"NBFC", "MUTHOOTFIN":"NBFC", "LTF":"NBFC",
    "PNBHOUSING":"NBFC", "LICHSGFIN":"NBFC",
    "TCS":"IT", "INFY":"IT", "HCLTECH":"IT", "WIPRO":"IT", "TECHM":"IT",
    "OFSS":"IT", "MPHASIS":"IT", "COFORGE":"IT", "PERSISTENT":"IT", "KPITTECH":"IT",
    "INDIGO":"AVIATION", "MOTHERSON":"AUTO", "MARUTI":"AUTO", "TVSMOTOR":"AUTO",
    "BAJAJ-AUTO":"AUTO", "HEROMOTOCO":"AUTO", "ASHOKLEY":"AUTO", "EICHERMOT":"AUTO",
    "HYUNDAI":"AUTO", "UNOMINDA":"AUTO",
    "SHREECEM":"CEMENT", "ULTRACEMCO":"CEMENT", "DALBHARAT":"CEMENT", "AMBUJACEM":"CEMENT",
    "TRENT":"RETAIL", "DMART":"RETAIL",
    "GODREJPROP":"REALTY", "LODHA":"REALTY", "DLF":"REALTY", "PRESTIGE":"REALTY",
    "OBEROIRLTY":"REALTY", "PHOENIXLTD":"REALTY",
    # Pharma
    "AUROPHARMA":"PHARMA", "CIPLA":"PHARMA", "SUNPHARMA":"PHARMA",
    "DRREDDY":"PHARMA", "LUPIN":"PHARMA", "DIVISLAB":"PHARMA",
    "ALKEM":"PHARMA", "TORNTPHARM":"PHARMA", "IPCALAB":"PHARMA",
    # Hospitals
    "MAXHEALTH":"HOSPITAL", "APOLLOHOSP":"HOSPITAL", "FORTIS":"HOSPITAL",
    # Hotels
    "INDHOTEL":"HOTEL", "LEMONTREE":"HOTEL", "CHALET":"HOTEL",
    # Power & infra
    "CGPOWER":"POWER", "NHPC":"POWER", "NTPC":"POWER", "POWERGRID":"POWER",
    "SIEMENS":"CAPITAL_GOODS", "ABB":"CAPITAL_GOODS", "BHEL":"CAPITAL_GOODS",
    # E-commerce / consumer
    "NYKAA":"ECOMM", "FIRSTCRY":"ECOMM", "ZOMATO":"ECOMM",
    # Auto ancillaries
    "SONACOMS":"AUTO_ANC", "BOSCHLTD":"AUTO_ANC", "BHARATFORG":"AUTO_ANC",
    "SUNDRMFAST":"AUTO_ANC",
    # Chemicals
    "PIDILITIND":"CHEMICALS", "AARTI":"CHEMICALS", "DEEPAKNTR":"CHEMICALS",
    # Telecom
    "BHARTIARTL":"TELECOM", "IDEA":"TELECOM",
    # FMCG
    "HINDUNILVR":"FMCG", "ITC":"FMCG", "NESTLEIND":"FMCG", "BRITANNIA":"FMCG",
    "DABUR":"FMCG", "MARICO":"FMCG", "COLPAL":"FMCG",
    # Insurance
    "HDFCLIFE":"INSURANCE", "SBILIFE":"INSURANCE", "ICICIPRULI":"INSURANCE",
    "LICI":"INSURANCE",
    # Oil & gas
    "RELIANCE":"OIL_GAS", "ONGC":"OIL_GAS", "BPCL":"OIL_GAS", "IOC":"OIL_GAS",
    "GAIL":"OIL_GAS",
    # Metals
    "TATASTEEL":"METALS", "HINDALCO":"METALS", "JSWSTEEL":"METALS",
    "SAIL":"METALS", "VEDL":"METALS", "NATIONALUM":"METALS",
    # Diversified
    "UNITDSPR":"DIVERSIFIED",
}


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# ── Portfolio persistence ─────────────────────────────────────────────────────
def _empty_portfolio():
    return {
        "total_capital":  TOTAL_CAPITAL,
        "cash":           TOTAL_CAPITAL,
        "positions":      {},
        "closed_trades":  [],
        "created":        datetime.now().isoformat(),
    }


def load_portfolio(path=PORTFOLIO_FILE):
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    log(f"  [portfolio] No existing portfolio at {path} — starting fresh with ₹10,00,000")
    return _empty_portfolio()


def save_portfolio(portfolio, path=PORTFOLIO_FILE):
    with open(path, "w") as f:
        json.dump(portfolio, f, indent=2)


# ── Price fetch (batched, shared Angel session) ─────────────────────────────────
def fetch_prices(symbols):
    """Fetch live LTP for NSE symbols."""
    if not symbols:
        return {}
    try:
        prices = fetch_ltps(symbols)
        return {sym: prices.get(sym) for sym in symbols}
    except Exception as e:
        log(f"  !! Angel One price fetch failed: {e}")
        return {s: None for s in symbols}


def _merge_prices(base, extra_symbols):
    """Fill missing symbols with one batched quote call."""
    merged = dict(base or {})
    missing = [s for s in extra_symbols if merged.get(s) is None]
    if missing:
        merged.update(fetch_prices(missing))
    return merged


# ── Scanner results ───────────────────────────────────────────────────────────
def load_signals():
    if not os.path.exists(SCAN_FILE):
        log(f"  !! {SCAN_FILE} not found — run long_buildup_scanner.py first")
        return [], "UNKNOWN", 0.0
    with open(SCAN_FILE) as f:
        data = json.load(f)
    macro    = data.get("macro", {}).get("sentiment", "UNKNOWN")
    fii_net  = float(data.get("macro", {}).get("fii_net_cr", 0) or 0)
    return data.get("results", []), macro, fii_net


# ── Excel report ──────────────────────────────────────────────────────────────
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREY_FILL  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HDR_FONT   = Font(bold=True)
CENTER     = Alignment(horizontal="center")


def _color_pnl_col(ws, col_letter):
    for row in ws.iter_rows(min_row=2, min_col=ws[col_letter + "1"].column,
                             max_col=ws[col_letter + "1"].column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.fill = GREEN_FILL if cell.value >= 0 else RED_FILL


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 30)


def _header_row(ws):
    for cell in ws[1]:
        cell.font = HDR_FONT
        cell.fill = GREY_FILL
        cell.alignment = CENTER


def _fmt_entry_val(val):
    if val is None:
        return "-"
    return val


def _model_liquidity_storage(record):
    """Liquidity + model score fields from scanner signal."""
    return {
        "ce_oi":            record.get("ce_oi"),
        "pe_oi":            record.get("pe_oi"),
        "total_oi":         record.get("total_oi"),
        "is_liquid":        record.get("is_liquid"),
        "quality_ratio":    record.get("quality_ratio"),
        "model_a":          record.get("model_a"),
        "model_b":          record.get("model_b"),
        "model_c":          record.get("model_c"),
        "model_d":          record.get("model_d"),
        "model_e":          record.get("model_e"),
        "model_f":          record.get("model_f"),
        "composite_abc":    record.get("composite_abc"),
        "model_f_pass":     record.get("model_f_pass"),
        "suggested_action": record.get("suggested_action"),
        "nifty_chg_pct":    record.get("nifty_chg_pct"),
    }


def _model_liquidity_excel_cols(record):
    liquid = record.get("is_liquid")
    if liquid is True:
        liquid_col = "Yes"
    elif liquid is False:
        liquid_col = "No"
    else:
        liquid_col = "-"
    mfp = record.get("model_f_pass")
    if mfp is True:
        mfp_col = "Yes"
    elif mfp is False:
        mfp_col = "No"
    else:
        mfp_col = "-"
    return {
        "CE OI":            _fmt_entry_val(record.get("ce_oi")),
        "PE OI":            _fmt_entry_val(record.get("pe_oi")),
        "Total OI":         _fmt_entry_val(record.get("total_oi")),
        "Liquid":           liquid_col,
        "Quality Ratio":    _fmt_entry_val(record.get("quality_ratio")),
        "Model A":          _fmt_entry_val(record.get("model_a")),
        "Model B":          _fmt_entry_val(record.get("model_b")),
        "Model C":          _fmt_entry_val(record.get("model_c")),
        "Model D":          _fmt_entry_val(record.get("model_d")),
        "Model E":          _fmt_entry_val(record.get("model_e")),
        "Model F":          _fmt_entry_val(record.get("model_f")),
        "Composite ABC":    _fmt_entry_val(record.get("composite_abc")),
        "Model F Pass":     mfp_col,
        "Suggested Action": record.get("suggested_action") or "-",
        "Nifty Chg %":      _fmt_entry_val(record.get("nifty_chg_pct")),
    }


_LEGACY_SIGNAL_COLS = {
    "Price Chg %": "-", "OI Chg %": "-", "Vol Ratio": "-", "PCR": "-",
    "Sector": "-", "EMA9": "-", "EMA21": "-", "EMA50": "-",
    "Latest OI": "-", "Entry Time": "-", "NSE TS": "-",
}
_LEGACY_MODEL_COLS = {
    "CE OI": "-", "PE OI": "-", "Total OI": "-", "Liquid": "-",
    "Quality Ratio": "-", "Model A": "-", "Model B": "-", "Model C": "-",
    "Model D": "-", "Model E": "-", "Model F": "-", "Composite ABC": "-",
    "Model F Pass": "-", "Suggested Action": "-", "Nifty Chg %": "-",
}


def _entry_storage(record):
    """Persist full scanner signal snapshot on positions and closed_trades."""
    return {
        # Signal quality fields
        "price_chg":   record.get("price_chg"),
        "oi_chg":      record.get("oi_chg"),
        "vol_ratio":   record.get("vol_ratio"),
        "pcr":         record.get("pcr"),
        "sector":      record.get("sector") or "-",
        "macro_entry": record.get("macro_entry", "UNKNOWN"),
        # EMA stack values at entry (for trend-strength analysis)
        "ema9":        record.get("ema9"),
        "ema21":       record.get("ema21"),
        "ema50":       record.get("ema50"),
        # Raw OI numbers (for absolute OI context, not just % change)
        "latest_oi":   record.get("latest_oi"),
        "prev_oi":     record.get("prev_oi"),
        # Entry time (so we can later analyse 3:15 vs 3:27 entries)
        "entry_time":  record.get("entry_time") or datetime.now().strftime("%H:%M:%S"),
        # NSE scan timestamp (tells us lag between scan data and entry)
        "nse_ts":      record.get("nse_ts"),
        **_model_liquidity_storage(record),
    }


def _has_entry_snapshot(record):
    """True if any entry-time signal field was stored on this trade."""
    return any(record.get(k) is not None for k in (
        "price_chg", "oi_chg", "vol_ratio", "pcr", "ema9", "latest_oi",
    ))


def _entry_excel_cols(record):
    """Map stored entry params to Excel columns.

    Legacy trades with no snapshot fields show '-'. Partial snapshots
    (e.g. pcr/oi_chg only) still render available columns.
    """
    macro = record.get("macro_entry", "?")
    if not _has_entry_snapshot(record):
        return {
            **_LEGACY_SIGNAL_COLS,
            "Macro": macro,
            **_LEGACY_MODEL_COLS,
        }
    return {
        "Price Chg %": _fmt_entry_val(record.get("price_chg")),
        "OI Chg %":    _fmt_entry_val(record.get("oi_chg")),
        "Vol Ratio":   _fmt_entry_val(record.get("vol_ratio")),
        "PCR":         _fmt_entry_val(record.get("pcr")),
        "Sector":      record.get("sector") or "-",
        "Macro":       macro,
        "EMA9":        _fmt_entry_val(record.get("ema9")),
        "EMA21":       _fmt_entry_val(record.get("ema21")),
        "EMA50":       _fmt_entry_val(record.get("ema50")),
        "Latest OI":   _fmt_entry_val(record.get("latest_oi")),
        "Entry Time":  record.get("entry_time") or "-",
        "NSE TS":      record.get("nse_ts") or "-",
        **_model_liquidity_excel_cols(record),
    }


ALL_TRADES_COLUMNS = [
    "Symbol", "Entry Date", "Exit Date", "Entry ₹", "Exit ₹", "CMP ₹", "Qty",
    "Live Trade", "Live Order ID",
    "Price Chg %", "OI Chg %", "Vol Ratio", "PCR", "Sector", "Macro",
    "EMA9", "EMA21", "EMA50", "Latest OI", "Entry Time", "NSE TS",
    "CE OI", "PE OI", "Total OI", "Liquid",
    "Quality Ratio", "Model A", "Model B", "Model C", "Model D", "Model E", "Model F",
    "Composite ABC", "Model F Pass", "Suggested Action", "Nifty Chg %",
    "P&L ₹", "P&L %", "Status", "Exit Reason",
]

EOD_CONFIRM_COLUMNS = [
    "Date", "Symbol", "Scan Time", "Confirm Time",
    "Price Chg %", "OI Chg %", "Vol 3:15", "Vol 3:27", "PCR", "Sector", "Macro",
    "CE OI", "PE OI", "Total OI", "Liquid",
    "Quality Ratio", "Model A", "Model B", "Model C", "Model D", "Model E", "Model F",
    "Composite ABC", "Model F Pass", "Suggested Action", "Nifty Chg %",
    "Open ₹", "High ₹", "Low ₹", "LTP ₹",
    "Close>Open", "Near High", "New High",
    "Confirm Pass", "Baseline Entered", "Confirm Entered", "Fail Reason",
]


def _bool_col(val):
    if val is True:
        return "Yes"
    if val is False:
        return "No"
    return "-"


def _load_confirm_log_entries():
    if not os.path.exists(CONFIRM_LOG_FILE):
        return []
    with open(CONFIRM_LOG_FILE) as f:
        data = json.load(f)
    return data.get("entries", []) if isinstance(data, dict) else []


def _trades_rows(portfolio, all_prices):
    """Build All Trades rows for a portfolio dict."""
    closed_rows, open_rows = [], []
    for t in portfolio.get("closed_trades", []):
        closed_rows.append({
            "Symbol": t["symbol"], "Entry Date": t["entry_date"],
            "Exit Date": t["exit_date"], "Entry ₹": t["entry_price"],
            "Exit ₹": t["exit_price"], "CMP ₹": t["exit_price"], "Qty": t["quantity"],
            "Live Trade": "Yes" if t.get("live_order_id") else "No",
            "Live Order ID": t.get("live_order_id") or "-",
            **_entry_excel_cols(t),
            "P&L ₹": t["pnl_abs"], "P&L %": t["pnl_pct"],
            "Status": "CLOSED", "Exit Reason": t["reason"],
        })
    for sym, pos in portfolio.get("positions", {}).items():
        curr = all_prices.get(sym) or pos["entry_price"]
        pnl = round((curr - pos["entry_price"]) * pos["quantity"], 2)
        pnl_pct = round((curr - pos["entry_price"]) / pos["entry_price"] * 100, 2)
        open_rows.append({
            "Symbol": sym, "Entry Date": pos["entry_date"],
            "Exit Date": "-", "Entry ₹": pos["entry_price"],
            "Exit ₹": "-", "CMP ₹": curr, "Qty": pos["quantity"],
            "Live Trade": "Yes" if pos.get("live_order_id") else "No",
            "Live Order ID": pos.get("live_order_id") or "-",
            **_entry_excel_cols(pos),
            "P&L ₹": pnl, "P&L %": pnl_pct,
            "Status": "OPEN", "Exit Reason": "-",
        })
    return open_rows + closed_rows


def _eod_confirm_rows():
    rows = []
    for e in _load_confirm_log_entries():
        rows.append({
            "Date": e.get("date"),
            "Symbol": e.get("symbol"),
            "Scan Time": e.get("scan_time", ""),
            "Confirm Time": e.get("confirm_time", ""),
            "Price Chg %": _fmt_entry_val(e.get("price_chg")),
            "OI Chg %": _fmt_entry_val(e.get("oi_chg")),
            "Vol 3:15": _fmt_entry_val(e.get("vol_ratio_scan")),
            "Vol 3:27": _fmt_entry_val(e.get("vol_ratio_confirm")),
            "PCR": _fmt_entry_val(e.get("pcr")),
            "Sector": e.get("sector", "-"),
            "Macro": e.get("macro", "?"),
            **_model_liquidity_excel_cols(e),
            "Open ₹": _fmt_entry_val(e.get("open")),
            "High ₹": _fmt_entry_val(e.get("high")),
            "Low ₹": _fmt_entry_val(e.get("low")),
            "LTP ₹": _fmt_entry_val(e.get("ltp")),
            "Close>Open": _bool_col(e.get("close_gt_open")),
            "Near High": _bool_col(e.get("near_day_high")),
            "New High": _bool_col(e.get("new_high")),
            "Confirm Pass": _bool_col(e.get("confirm_pass")),
            "Baseline Entered": _bool_col(e.get("baseline_entered")),
            "Confirm Entered": _bool_col(e.get("confirm_entered")),
            "Fail Reason": e.get("fail_reason") or "-",
        })
    return rows


PNL_COLUMNS = ["Date", "Trades", "P&L ₹", "Wins", "Losses", "Win Rate %", "Cumul P&L ₹"]
PNL_WEEK_COLUMNS = ["Week", "Trades", "P&L ₹", "Wins", "Losses", "Win Rate %", "Cumul P&L ₹"]
PNL_MONTH_COLUMNS = ["Month", "Trades", "P&L ₹", "Wins", "Losses", "Win Rate %", "Cumul P&L ₹"]


def _build_daily_pnl(trades):
    if not trades:
        return pd.DataFrame(columns=PNL_COLUMNS)
    df2 = pd.DataFrame(trades)
    daily = (df2.groupby("exit_date")
                .agg(Trades=("symbol", "count"),
                     PnL=("pnl_abs", "sum"),
                     Wins=("pnl_abs", lambda x: (x > 0).sum()),
                     Losses=("pnl_abs", lambda x: (x <= 0).sum()))
                .reset_index()
                .sort_values("exit_date"))
    daily.columns = ["Date", "Trades", "P&L ₹", "Wins", "Losses"]
    daily["Win Rate %"] = (daily["Wins"] / daily["Trades"] * 100).round(1)
    daily["P&L ₹"] = daily["P&L ₹"].round(2)
    daily["Cumul P&L ₹"] = daily["P&L ₹"].cumsum().round(2)
    return daily


def _build_weekly_pnl(trades):
    if not trades:
        return pd.DataFrame(columns=PNL_WEEK_COLUMNS)
    df3 = pd.DataFrame(trades)
    df3["exit_date"] = pd.to_datetime(df3["exit_date"])
    df3["Week"] = df3["exit_date"].dt.strftime("W%V %Y")
    df3["wk_sort"] = df3["exit_date"].dt.strftime("%Y-W%V")
    weekly = (df3.groupby(["wk_sort", "Week"])
                 .agg(Trades=("symbol", "count"),
                      PnL=("pnl_abs", "sum"),
                      Wins=("pnl_abs", lambda x: (x > 0).sum()),
                      Losses=("pnl_abs", lambda x: (x <= 0).sum()))
                 .reset_index()
                 .sort_values("wk_sort")
                 .drop("wk_sort", axis=1))
    weekly.columns = ["Week", "Trades", "P&L ₹", "Wins", "Losses"]
    weekly["Win Rate %"] = (weekly["Wins"] / weekly["Trades"] * 100).round(1)
    weekly["P&L ₹"] = weekly["P&L ₹"].round(2)
    weekly["Cumul P&L ₹"] = weekly["P&L ₹"].cumsum().round(2)
    return weekly


def _build_monthly_pnl(trades):
    if not trades:
        return pd.DataFrame(columns=PNL_MONTH_COLUMNS)
    df4 = pd.DataFrame(trades)
    df4["exit_date"] = pd.to_datetime(df4["exit_date"])
    df4["Month"] = df4["exit_date"].dt.strftime("%b %Y")
    df4["mo_sort"] = df4["exit_date"].dt.strftime("%Y-%m")
    monthly = (df4.groupby(["mo_sort", "Month"])
                  .agg(Trades=("symbol", "count"),
                       PnL=("pnl_abs", "sum"),
                       Wins=("pnl_abs", lambda x: (x > 0).sum()),
                       Losses=("pnl_abs", lambda x: (x <= 0).sum()))
                  .reset_index()
                  .sort_values("mo_sort")
                  .drop("mo_sort", axis=1))
    monthly.columns = ["Month", "Trades", "P&L ₹", "Wins", "Losses"]
    monthly["Win Rate %"] = (monthly["Wins"] / monthly["Trades"] * 100).round(1)
    monthly["P&L ₹"] = monthly["P&L ₹"].round(2)
    monthly["Cumul P&L ₹"] = monthly["P&L ₹"].cumsum().round(2)
    return monthly


def _open_positions_rows(portfolio, all_prices):
    rows = []
    for sym, pos in portfolio.get("positions", {}).items():
        curr = all_prices.get(sym) or pos["entry_price"]
        pnl = round((curr - pos["entry_price"]) * pos["quantity"], 2)
        pnl_pct = round((curr - pos["entry_price"]) / pos["entry_price"] * 100, 2)
        rows.append({
            "Symbol": sym,
            "Entry Date": pos["entry_date"],
            "Entry ₹": pos["entry_price"],
            "Live Trade": "Yes" if pos.get("live_order_id") else "No",
            "Live Order ID": pos.get("live_order_id") or "-",
            **_entry_excel_cols(pos),
            "CMP ₹": curr,
            "Qty": pos["quantity"],
            "Invested ₹": pos["invested"],
            "Unrealised ₹": pnl,
            "P&L %": pnl_pct,
            "SL ₹": pos["stop_loss"],
            "Target ₹": pos["target"],
        })
    open_df = pd.DataFrame(rows)
    if not open_df.empty:
        open_df["_sort"] = pd.to_datetime(open_df["Entry Date"], errors="coerce")
        open_df = open_df.sort_values("_sort", ascending=False).drop(columns="_sort").reset_index(drop=True)
    return open_df


def generate_excel(portfolio, all_prices=None, confirm_portfolio=None):
    trades = portfolio.get("closed_trades", [])
    if confirm_portfolio is None:
        confirm_portfolio = load_portfolio(CONFIRM_PORTFOLIO_FILE)
    confirm_trades = confirm_portfolio.get("closed_trades", [])

    syms = list(set(portfolio.get("positions", {})) | set(confirm_portfolio.get("positions", {})))
    if all_prices is None:
        all_prices = fetch_prices(syms) if syms else {}
    else:
        all_prices = _merge_prices(all_prices, syms)

    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:

        # ── Sheet 1: All Trades (baseline) ───────────────────────────────────
        all_rows = _trades_rows(portfolio, all_prices)
        df = pd.DataFrame(all_rows, columns=ALL_TRADES_COLUMNS)
        if not df.empty:
            df["_sort"] = pd.to_datetime(df["Entry Date"], errors="coerce")
            df = df.sort_values("_sort", ascending=False).drop(columns="_sort").reset_index(drop=True)
        df.to_excel(writer, sheet_name="All Trades", index=False)

        # ── Sheet 2: Confirm Trades (shadow strategy) ────────────────────────
        confirm_rows = _trades_rows(confirm_portfolio, all_prices)
        cdf = pd.DataFrame(confirm_rows, columns=ALL_TRADES_COLUMNS)
        if not cdf.empty:
            cdf["_sort"] = pd.to_datetime(cdf["Entry Date"], errors="coerce")
            cdf = cdf.sort_values("_sort", ascending=False).drop(columns="_sort").reset_index(drop=True)
        cdf.to_excel(writer, sheet_name="Confirm Trades", index=False)

        # ── Sheet 3: EOD Confirm log ─────────────────────────────────────────
        eod_rows = _eod_confirm_rows()
        edf = pd.DataFrame(eod_rows, columns=EOD_CONFIRM_COLUMNS)
        if not edf.empty:
            edf["_sort"] = pd.to_datetime(edf["Date"], errors="coerce")
            edf = edf.sort_values("_sort", ascending=False).drop(columns="_sort").reset_index(drop=True)
        edf.to_excel(writer, sheet_name="EOD Confirm", index=False)

        # ── Sheet 4: Open Positions (baseline) ────────────────────────────────
        open_df = _open_positions_rows(portfolio, all_prices)
        open_df.to_excel(writer, sheet_name="Open Positions", index=False)

        # ── Sheet 5: Confirm Open Positions ───────────────────────────────────
        confirm_open_df = _open_positions_rows(confirm_portfolio, all_prices)
        confirm_open_df.to_excel(writer, sheet_name="Confirm Open Positions", index=False)

        # ── Sheet 6–8: Baseline P&L ───────────────────────────────────────────
        _build_daily_pnl(trades).to_excel(writer, sheet_name="Daily P&L", index=False)
        _build_weekly_pnl(trades).to_excel(writer, sheet_name="Weekly P&L", index=False)
        _build_monthly_pnl(trades).to_excel(writer, sheet_name="Monthly P&L", index=False)

        # ── Sheet 9–11: Confirm strategy P&L ──────────────────────────────────
        _build_daily_pnl(confirm_trades).to_excel(writer, sheet_name="Confirm Daily P&L", index=False)
        _build_weekly_pnl(confirm_trades).to_excel(writer, sheet_name="Confirm Weekly P&L", index=False)
        _build_monthly_pnl(confirm_trades).to_excel(writer, sheet_name="Confirm Monthly P&L", index=False)

    # ── Apply colour + formatting ─────────────────────────────────────────────
    wb = load_workbook(EXCEL_FILE)

    sheet_names = [
        "All Trades", "Confirm Trades", "EOD Confirm",
        "Open Positions", "Confirm Open Positions",
        "Daily P&L", "Weekly P&L", "Monthly P&L",
        "Confirm Daily P&L", "Confirm Weekly P&L", "Confirm Monthly P&L",
    ]
    for sname in sheet_names:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        _header_row(ws)
        _autofit(ws)
        for cell in ws[1]:
            if cell.value and "P&L ₹" in str(cell.value) and "Cumul" not in str(cell.value):
                _color_pnl_col(ws, get_column_letter(cell.column))
        if sname == "EOD Confirm":
            pass_col = None
            for cell in ws[1]:
                if cell.value == "Confirm Pass":
                    pass_col = get_column_letter(cell.column)
                    break
            if pass_col:
                for row in ws.iter_rows(min_row=2, min_col=ws[f"{pass_col}1"].column,
                                        max_col=ws[f"{pass_col}1"].column):
                    for cell in row:
                        if cell.value == "Yes":
                            cell.fill = GREEN_FILL
                        elif cell.value == "No":
                            cell.fill = RED_FILL

    wb.save(EXCEL_FILE)
    log(f"  Trade log saved → {EXCEL_FILE}  (sheets: {' | '.join(sheet_names)})")


# ── Exit / entry helpers ──────────────────────────────────────────────────────
def _process_exits(portfolio, curr_prices, today, label=""):
    """Check SL/target on open positions; return (portfolio, closed_count)."""
    prefix = f"[{label}] " if label else ""
    to_close = []

    for sym, pos in list(portfolio["positions"].items()):
        curr = curr_prices.get(sym)
        if curr is None:
            log(f"  ? {prefix}{sym:<15} — price unavailable, skipping exit check")
            continue

        pnl_pct = round((curr - pos["entry_price"]) / pos["entry_price"] * 100, 2)
        pnl_abs = round((curr - pos["entry_price"]) * pos["quantity"], 2)

        reason = None
        if curr <= pos["stop_loss"]:
            reason = "SL HIT"
        elif curr >= pos["target"]:
            reason = "TARGET HIT"

        sign = "+" if pnl_abs >= 0 else ""
        marker = "✗" if reason else " "
        log(f"  {marker} {prefix}{sym:<15} CMP ₹{curr:>8,.2f}  Entry ₹{pos['entry_price']:>8,.2f}"
            f"  {sign}{pnl_pct:.2f}%  ₹{sign}{pnl_abs:>8,.0f}"
            f"{'  → ' + reason if reason else ''}")

        if reason:
            to_close.append((sym, curr, pnl_abs, pnl_pct, reason))

    for sym, exit_px, pnl_abs, pnl_pct, reason in to_close:
        pos = portfolio["positions"].pop(sym)
        entry_dt = pos.get("entry_date", today)
        try:
            hold_days = (datetime.strptime(today, "%Y-%m-%d") - datetime.strptime(entry_dt, "%Y-%m-%d")).days
        except Exception:
            hold_days = None
        portfolio["cash"] += round(exit_px * pos["quantity"], 2)
        portfolio["closed_trades"].append({
            "symbol": sym,
            "entry_date": pos["entry_date"],
            "exit_date": today,
            "entry_price": pos["entry_price"],
            "exit_price": exit_px,
            "quantity": pos["quantity"],
            "pnl_abs": pnl_abs,
            "pnl_pct": pnl_pct,
            "reason": reason,
            "hold_days": hold_days,
            "live_order_id": pos.get("live_order_id"),
            **_entry_storage(pos),
        })
        sign = "+" if pnl_abs >= 0 else ""
        log(f"  ✗ {prefix}CLOSED {sym:<14} Exit ₹{exit_px:.2f}  P&L ₹{sign}{pnl_abs:,.0f} ({sign}{pnl_pct:.2f}%)  [{reason}]")
        # NOTE: any live order for this symbol is NOT closed here — the paper
        # SL(-1%)/target(+2%) differs from the live trade's own SL(-0.65%)/
        # target(+2%). Live positions are monitored independently via
        # LIVE_TRADER.check_and_exit_positions() in run().

        try:
            import bot_status
            bot_status.log_activity("trade", f"CLOSED {sym} [{reason}] P&L ₹{sign}{pnl_abs:,.0f}",
                                    symbol=sym, pnl=pnl_abs, reason=reason)
        except Exception:
            pass

    return portfolio, len(to_close)


def _open_positions(portfolio, signals, macro, fii_net, today, label="", entry_prices=None):
    """Open new positions from scanner signals. Returns set of symbols entered."""
    already_open = set(portfolio["positions"].keys())
    from datetime import timedelta as _td
    _cutoff = (datetime.now() - _td(days=3)).strftime("%Y-%m-%d")
    recently_traded = {
        t["symbol"] for t in portfolio["closed_trades"]
        if t.get("exit_date", "") >= _cutoff
    }
    new_signals = [r for r in signals if r["symbol"] not in already_open and r["symbol"] not in recently_traded]
    slots_free = MAX_POSITIONS - len(portfolio["positions"])

    log(f"\n  {label}Signals: {len(signals)}  |  New: {len(new_signals)}  |  Slots: {slots_free}  |  Cash: ₹{portfolio['cash']:,.0f}")

    if macro == "BEARISH":
        log(f"  {label}*** MACRO BEARISH — no new entries today ***")
        return set()

    fii_selling = fii_net < -500
    if fii_selling:
        log(f"  {label}⚠  FII net selling (₹{fii_net:+,.0f} Cr) — BANK/NBFC blocked")

    sector_count = {}
    for sym_open in portfolio["positions"]:
        sec = SECTOR_MAP.get(sym_open)
        if sec:
            sector_count[sec] = sector_count.get(sec, 0) + 1

    if entry_prices is None and new_signals:
        entry_prices = fetch_prices([r["symbol"] for r in new_signals])
    entry_prices = entry_prices or {}

    entered = set()
    count = 0
    for sig in new_signals:
        if count >= slots_free:
            break
        sym = sig["symbol"]
        scan_px = sig.get("spot_price") or 0
        entry_px = entry_prices.get(sym) or scan_px
        if not entry_px or entry_px <= 0:
            continue
        entry_px = round(float(entry_px), 2)
        if scan_px and abs(entry_px - scan_px) >= 0.05:
            log(f"  {label}  {sym}: scanner ₹{scan_px:,.2f} → live entry ₹{entry_px:,.2f}")

        sec = SECTOR_MAP.get(sym)
        if sec and sector_count.get(sec, 0) >= MAX_PER_SECTOR:
            log(f"  {label}— SKIP  {sym:<14} sector {sec} already has {sector_count[sec]} position(s)")
            continue
        if fii_selling and sec in ("BANK", "NBFC"):
            log(f"  {label}— SKIP  {sym:<14} FII selling — {sec} blocked")
            continue

        allocated = min(ALLOC_PER_TRADE, portfolio["cash"])
        if allocated < entry_px:
            log(f"  {label}!! {sym}: not enough cash (need ₹{entry_px:.2f}, have ₹{portfolio['cash']:,.0f})")
            continue
        qty = int(allocated // entry_px)
        if qty == 0:
            continue
        invested = round(qty * entry_px, 2)
        sl_px = round(entry_px * (1 - STOP_LOSS_PCT / 100), 2)
        tgt_px = round(entry_px * (1 + TARGET_PCT / 100), 2)

        portfolio["cash"] -= invested
        portfolio["positions"][sym] = {
            "entry_price": entry_px,
            "quantity": qty,
            "invested": invested,
            "entry_date": today,
            "stop_loss": sl_px,
            "target": tgt_px,
            **_entry_storage({
                **sig,
                "sector": sec or "-",
                "macro_entry": macro,
            }),
        }
        if sec:
            sector_count[sec] = sector_count.get(sec, 0) + 1
        count += 1
        entered.add(sym)
        log(f"  {label}+ OPEN  {sym:<14} {qty} sh @ ₹{entry_px:,.2f}"
            f"  invested ₹{invested:,.0f}  SL ₹{sl_px:.2f}  T ₹{tgt_px:.2f}  [{macro}]")

        # Attempt live order if conditions met (Model B >= 90 AND vol_ratio >= 1.5).
        # Live sizing/SL/target are independent of the paper trade above —
        # see live_trader.py — so only the order_id is tagged here for reporting.
        if LIVE_TRADER:
            try:
                live_result = LIVE_TRADER.place_live_order(sig, entry_px)
                if live_result["placed"]:
                    log(f"  {label}🔴 LIVE  {sym:<14} Order ID: {live_result['order_id']}")
                    portfolio["positions"][sym]["live_order_id"] = live_result["order_id"]
                else:
                    log(f"  {label}⚪ PAPER-ONLY {sym:<14} ({live_result['reason']})")
            except Exception as e:
                log(f"  {label}⚠ Live order error for {sym}: {e}")

        try:
            import bot_status
            bot_status.log_activity("trade", f"OPEN {sym} {qty}sh @ ₹{entry_px:,.2f} [{macro}]",
                                    symbol=sym, qty=qty, price=entry_px)
        except Exception:
            pass

    return entered


def enter_confirm_positions(signals, macro, fii_net, entry_prices=None):
    """Open shadow positions for EOD Confirm strategy (called at 3:30)."""
    portfolio = load_portfolio(CONFIRM_PORTFOLIO_FILE)
    today = datetime.now().strftime("%Y-%m-%d")
    log("\n  ── Confirm Strategy Entries (3:30) ──")
    entered = _open_positions(
        portfolio, signals, macro, fii_net, today,
        label="[confirm] ", entry_prices=entry_prices,
    )
    save_portfolio(portfolio, CONFIRM_PORTFOLIO_FILE)
    return entered


# ── Main ──────────────────────────────────────────────────────────────────────
def run(intraday_only=False):
    portfolio = load_portfolio()
    confirm_pf = load_portfolio(CONFIRM_PORTFOLIO_FILE)
    signals, macro, fii_net = load_signals()
    today = datetime.now().strftime("%Y-%m-%d")

    log("=" * 72)
    log("  PAPER TRADER — ₹10,00,000 Virtual Portfolio")
    log(f"  Date       : {datetime.now().strftime('%d %b %Y %H:%M IST')}")
    log(f"  Cash       : ₹{portfolio['cash']:>10,.0f}  |  Open: {len(portfolio['positions'])} positions")
    log(f"  Confirm    : ₹{confirm_pf['cash']:>10,.0f}  |  Open: {len(confirm_pf['positions'])} positions")
    log(f"  FII Macro  : {macro}  |  SL: -{STOP_LOSS_PCT}%  Target: +{TARGET_PCT}%  Alloc: ₹{ALLOC_PER_TRADE:,.0f}/trade")
    log("=" * 72)

    all_syms = list(set(portfolio["positions"]) | set(confirm_pf["positions"]))
    curr_prices = fetch_prices(all_syms) if all_syms else {}

    # Live positions monitored on their OWN SL(-0.65%)/target(+2%) — independent
    # of the paper portfolio's -1%/+2% rule above. Additive only; does not
    # touch paper_portfolio.json or confirm portfolio.
    if LIVE_TRADER:
        try:
            live_syms = LIVE_TRADER.get_live_symbols()
            missing_live = [s for s in live_syms if s not in curr_prices]
            if missing_live:
                curr_prices.update(fetch_prices(missing_live))
            LIVE_TRADER.check_and_exit_positions(curr_prices)
        except Exception as e:
            log(f"  ⚠ Live position monitor error: {e}")

    log("\n  Checking open positions (baseline)...")
    if not portfolio["positions"]:
        log("  (no open baseline positions)")
    portfolio, _ = _process_exits(portfolio, curr_prices, today)

    log("\n  Checking open positions (confirm)...")
    if not confirm_pf["positions"]:
        log("  (no open confirm positions)")
    confirm_pf, _ = _process_exits(confirm_pf, curr_prices, today, label="confirm")

    if intraday_only:
        log("  (intraday mode — skipping new entries, SL/target check only)")
        save_portfolio(portfolio)
        save_portfolio(confirm_pf, CONFIRM_PORTFOLIO_FILE)
        all_prices = curr_prices
        generate_excel(portfolio, all_prices, confirm_pf)
        log("=" * 72)
        return

    if macro == "CAUTIOUS":
        log("  **  MACRO CAUTIOUS — FII net selling. Entering with reduced conviction. **")

    from datetime import timedelta as _td
    _cutoff = (datetime.now() - _td(days=3)).strftime("%Y-%m-%d")
    recently_traded = {
        t["symbol"] for t in portfolio["closed_trades"]
        if t.get("exit_date", "") >= _cutoff
    }
    skipped_recent = [r["symbol"] for r in signals if r["symbol"] in recently_traded]
    if skipped_recent:
        log(f"  Skipped (traded within 3 days): {', '.join(skipped_recent)}")

    log("\n  ── Baseline Entries (3:25) ──")
    _open_positions(portfolio, signals, macro, fii_net, today)

    all_syms = list(set(portfolio["positions"]) | set(confirm_pf["positions"]))
    all_prices = _merge_prices(curr_prices, all_syms)

    unrealised  = sum(
        ((all_prices.get(s) or p["entry_price"]) - p["entry_price"]) * p["quantity"]
        for s, p in portfolio["positions"].items()
    )
    invested_val = sum(p["invested"] for p in portfolio["positions"].values())
    portfolio_val = portfolio["cash"] + invested_val + unrealised
    total_pnl     = portfolio_val - TOTAL_CAPITAL
    total_pnl_pct = round(total_pnl / TOTAL_CAPITAL * 100, 2)
    closed_pnl    = sum(t["pnl_abs"] for t in portfolio["closed_trades"])
    wins          = [t for t in portfolio["closed_trades"] if t["pnl_abs"] > 0]
    losses        = [t for t in portfolio["closed_trades"] if t["pnl_abs"] <= 0]
    win_rate      = round(len(wins) / len(portfolio["closed_trades"]) * 100) if portfolio["closed_trades"] else 0

    # ── Step 5: Print open positions ──────────────────────────────────────────
    if portfolio["positions"]:
        log("\n  Open Positions:")
        print(f"\n  {'Symbol':<15} {'Entry ₹':>9} {'CMP ₹':>9} {'Qty':>6}"
              f" {'SL ₹':>9} {'Target ₹':>10} {'P&L ₹':>10} {'P&L%':>7}  Macro")
        print("  " + "-" * 90)
        for sym, pos in portfolio["positions"].items():
            curr    = all_prices.get(sym) or pos["entry_price"]
            pnl     = round((curr - pos["entry_price"]) * pos["quantity"], 2)
            pnl_pct = round((curr - pos["entry_price"]) / pos["entry_price"] * 100, 2)
            sign    = "+" if pnl >= 0 else ""
            print(f"  {sym:<15} {pos['entry_price']:>9.2f} {curr:>9.2f} {pos['quantity']:>6}"
                  f" {pos['stop_loss']:>9.2f} {pos['target']:>10.2f}"
                  f" {sign}{pnl:>9,.0f} {sign}{pnl_pct:>6.2f}%  {pos.get('macro_entry','?')}")
        print()

    # ── Step 6: Closed trades table ───────────────────────────────────────────
    if portfolio["closed_trades"]:
        log("  Closed Trades:")
        print(f"\n  {'Symbol':<15} {'Entry':>10} {'Exit':>10} {'Qty':>6}"
              f" {'P&L ₹':>11} {'P&L%':>8}  Reason")
        print("  " + "-" * 72)
        for t in sorted(portfolio["closed_trades"], key=lambda x: x["exit_date"], reverse=True):
            sign = "+" if t["pnl_abs"] >= 0 else ""
            print(f"  {t['symbol']:<15} {t['entry_price']:>10.2f} {t['exit_price']:>10.2f}"
                  f" {t['quantity']:>6} {sign}{t['pnl_abs']:>10,.0f} {sign}{t['pnl_pct']:>7.2f}%  {t['reason']}")
        print()

    # ── Step 7: Summary ───────────────────────────────────────────────────────
    sign = "+" if total_pnl >= 0 else ""
    log("=" * 72)
    log("  SUMMARY")
    log(f"  Portfolio Value : ₹{portfolio_val:>12,.0f}   ({sign}₹{total_pnl:,.0f}  {sign}{total_pnl_pct:.2f}%)")
    log(f"  Cash Available  : ₹{portfolio['cash']:>12,.0f}")
    log(f"  Invested        : ₹{invested_val:>12,.0f}   ({len(portfolio['positions'])} positions open)")
    log(f"  Unrealised P&L  : ₹{'+' if unrealised >= 0 else ''}{unrealised:>11,.0f}")
    log(f"  Realised P&L    : ₹{'+' if closed_pnl >= 0 else ''}{closed_pnl:>11,.0f}   ({len(wins)}W / {len(losses)}L  win rate {win_rate}%)")
    log("=" * 72)

    save_portfolio(portfolio)
    save_portfolio(confirm_pf, CONFIRM_PORTFOLIO_FILE)
    generate_excel(portfolio, all_prices, confirm_pf)
    log("=" * 72)


def refresh_excel():
    """Regenerate trade_log.xlsx from current portfolio state."""
    portfolio = load_portfolio()
    confirm_pf = load_portfolio(CONFIRM_PORTFOLIO_FILE)
    syms = list(set(portfolio["positions"]) | set(confirm_pf["positions"]))
    prices = fetch_prices(syms) if syms else {}
    generate_excel(portfolio, prices, confirm_pf)


if __name__ == "__main__":
    run()
